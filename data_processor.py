"""
Data Processor for Invoice Extraction System
Supports BOTH Agency Invoices (GroupM/M-Six) and Broadcaster Invoices.

Agency Sheet   — Channel-wise aggregated data from agency PDFs
Broadcaster Sheet — Individual spot data from broadcaster PDFs with 18 fields
"""

import os
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict
import logging

from pdf_parser import ParsedInvoice, parse_invoice, _parse_number
from broadcaster_types import ParsedBroadcasterInvoice

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────
# SHEET 1 — Invoice Summary (one row per invoice)
# ─────────────────────────────────────────────────────────────────

SUMMARY_COLUMNS = [
    'Invoice Number',
    'Invoice Date',
    'Agency Name',
    'Advertiser Name',
    'Brand Name',
    'Campaign Name',
    'Campaign Period (Activity Month)',
    'Estimate Number',
    'Estimate Period',
    'PO Number',
    'Total Spots',
    'Total Net Cost (₹)',
    'Total Amount Payable (₹)',
]


def build_invoice_summary(parsed_invoices: List[ParsedInvoice]) -> pd.DataFrame:
    rows = []
    for r in parsed_invoices:
        h = r.header
        rows.append({
            'Invoice Number':                   h.invoice_number,
            'Invoice Date':                     h.invoice_date,
            'Agency Name':                      h.agency_name,
            'Advertiser Name':                  h.advertiser_name,
            'Brand Name':                       h.brand_name,
            'Campaign Name':                    h.campaign_name,
            'Campaign Period (Activity Month)': h.activity_month,
            'Estimate Number':                  h.estimate_number,
            'Estimate Period':                  h.estimate_period,
            'PO Number':                        h.po_number,
            'Total Spots':                      r.total_spots if r.total_spots > 0 else len(r.spot_details),
            'Total Net Cost (₹)':               _parse_number(h.net_cost),
            'Total Amount Payable (₹)':         _parse_number(h.total_amount_payable),
        })
    df = pd.DataFrame(rows, columns=SUMMARY_COLUMNS)
    return df


# ─────────────────────────────────────────────────────────────────
# SHEET 2 — Channel-wise Aggregated Data
# Groups spots by: Invoice × Channel × Program × Time Band × Producer × Spot Rate
# Shows: No. of Spots, Total Net Cost, Date-wise Spots
# ─────────────────────────────────────────────────────────────────

CHANNEL_COLUMNS = [
    'Agency Name',
    'Advertiser Name',
    'Invoice Number',
    'Invoice Date',
    'Campaign Period (Activity Month)',
    'Estimate Number',
    'Estimate Period',
    'PO Number',
    'Brand Name',
    'Campaign Name',
    'Total Value Including Taxes (Rs)',
    'Channel Name',
    'Program',
    'Time Band',
    'Broadcaster Name (Producer)',
    'Spot Duration (Sec)',
    'Spot Rate (Per 10 Sec) Rs',
    'No. of Spots',
    'Total Net Cost (Rs)',
    'Date Wise Spots',
]


def build_channel_wise_data(parsed_invoices: List[ParsedInvoice]) -> pd.DataFrame:
    """
    Aggregate spot-level data into channel+program+rate groups.
    Each row = unique combination of Invoice + Channel + Program + Rate.
    Much easier to read than 11,000 individual spot rows.
    """
    rows = []

    for result in parsed_invoices:
        h = result.header

        # Build Annexure-1 lookup: (channel, program) -> date_wise_spots
        ann1_lookup: Dict[tuple, str] = {}
        for ar in result.annexure_records:
            key = (ar.channel.strip(), ar.program.strip())
            ann1_lookup[key] = ar.dates_with_spots

        # Build fallback date lookup from individual spot records
        spot_dates_lookup: Dict[tuple, set] = {}
        for sd in result.spot_details:
            key = (sd.channel.strip(), sd.program.strip())
            if key not in spot_dates_lookup:
                spot_dates_lookup[key] = set()
            if sd.spot_date:
                spot_dates_lookup[key].add(sd.spot_date)

        # Group spots by (channel, program, time_band, producer, duration, rate)
        groups: Dict[tuple, dict] = {}
        for sd in result.spot_details:
            key = (
                sd.channel.strip(),
                sd.program.strip(),
                sd.time_band.strip(),
                sd.producer.strip(),
                sd.spot_duration,
                sd.net_spot_rate_per_10sec,
            )
            if key not in groups:
                groups[key] = {'count': 0, 'net_cost': 0.0}
            groups[key]['count']    += 1
            groups[key]['net_cost'] += sd.net_cost

        for key, agg in groups.items():
            channel, program, time_band, producer, duration, rate = key

            # Use Annexure-1 date-wise spots; fallback to sorted individual dates
            date_wise = ann1_lookup.get((channel, program), '')
            if not date_wise:
                dates = sorted(spot_dates_lookup.get((channel, program), set()))
                date_wise = ', '.join(dates) if dates else ''

            rows.append({
                'Agency Name':                      h.agency_name,
                'Advertiser Name':                  h.advertiser_name,
                'Invoice Number':                   h.invoice_number,
                'Invoice Date':                     h.invoice_date,
                'Campaign Period (Activity Month)': h.activity_month,
                'Estimate Number':                  h.estimate_number,
                'Estimate Period':                  h.estimate_period,
                'PO Number':                        h.po_number,
                'Brand Name':                       h.brand_name,
                'Campaign Name':                    h.campaign_name,
                'Total Value Including Taxes (Rs)': _parse_number(h.total_amount_payable),
                'Channel Name':                     channel,
                'Program':                          program,
                'Time Band':                        time_band,
                'Broadcaster Name (Producer)':      producer,
                'Spot Duration (Sec)':              duration,
                'Spot Rate (Per 10 Sec) Rs':        rate,
                'No. of Spots':                     agg['count'],
                'Total Net Cost (Rs)':              round(agg['net_cost'], 2),
                'Date Wise Spots':                  date_wise,
            })

    if not rows:
        return pd.DataFrame(columns=CHANNEL_COLUMNS)

    df = pd.DataFrame(rows, columns=CHANNEL_COLUMNS)
    df = df.sort_values(['Invoice Number', 'Channel Name', 'Program']).reset_index(drop=True)
    return df


def build_channel_summary(parsed_invoices: List[ParsedInvoice]) -> pd.DataFrame:
    """
    Groups all spots across all invoices by Channel Name.
    Shows: Channel Name, Total Spots, Total Net Cost (Rs), Share of Budget (%)
    Very high-level and readable (10-15 rows).
    """
    channel_data = {}
    total_campaign_cost = 0.0
    
    for result in parsed_invoices:
        for sd in result.spot_details:
            channel = sd.channel.strip()
            if not channel:
                continue
            if channel not in channel_data:
                channel_data[channel] = {'spots': 0, 'cost': 0.0}
            channel_data[channel]['spots'] += 1
            channel_data[channel]['cost'] += sd.net_cost
            total_campaign_cost += sd.net_cost
            
    rows = []
    for channel, agg in channel_data.items():
        share = (agg['cost'] / total_campaign_cost * 100) if total_campaign_cost > 0 else 0.0
        rows.append({
            'Channel Name': channel,
            'Total Spots': agg['spots'],
            'Total Net Cost (Rs)': round(agg['cost'], 2),
            'Share of Budget (%)': round(share, 2)
        })
        
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values('Total Net Cost (Rs)', ascending=False).reset_index(drop=True)
    return df


# ─────────────────────────────────────────────────────────────────
# Process + Export
# ─────────────────────────────────────────────────────────────────

def process_multiple_invoices(parsed_invoices: List[ParsedInvoice]) -> Dict[str, pd.DataFrame]:
    return {
        'summary':         build_invoice_summary(parsed_invoices),
        'channel_summary': build_channel_summary(parsed_invoices),
        'channel':         build_channel_wise_data(parsed_invoices),
        # Keep raw flat sheet for internal use (not exported to main Excel)
        'main':            _build_flat_sheet(parsed_invoices),
    }


def _build_flat_sheet(parsed_invoices: List[ParsedInvoice]) -> pd.DataFrame:
    """Internal flat sheet (kept for app.py preview, not in main Excel)."""
    rows = []
    for result in parsed_invoices:
        h = result.header
        ann1: Dict[tuple, str] = {}
        for ar in result.annexure_records:
            ann1[(ar.channel.strip(), ar.program.strip())] = ar.dates_with_spots
        for sd in result.spot_details:
            rows.append({
                'Invoice Number':                   sd.invoice_number or h.invoice_number,
                'Channel Name':                     sd.channel,
                'Program':                          sd.program,
                'Time Band':                        sd.time_band,
                'Broadcaster Name (Producer)':      sd.producer,
                'Date':                             sd.spot_date,
                'Date Wise Spots':                  ann1.get((sd.channel.strip(), sd.program.strip()), ''),
                'Spot Duration (Sec)':              sd.spot_duration,
                'Spot Rate (Per 10 Sec)':           sd.net_spot_rate_per_10sec,
                'Net Cost':                         sd.net_cost,
            })
    return pd.DataFrame(rows) if rows else pd.DataFrame()


# ─────────────────────────────────────────────────────────────────
# Excel Styling
# ─────────────────────────────────────────────────────────────────

COLORS = {
    'sheet1_header': '1F3864',   # dark navy
    'sheet2_header': '5B2C6F',   # premium violet/purple
    'sheet3_header': '1A5276',   # dark teal
    'row_even':      'EBF5FB',
    'row_odd':       'FFFFFF',
    'total_row':     'D4E6F1',
}


def _style_header(ws, num_cols: int, color: str):
    hfont  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    hfill  = PatternFill(start_color=color, end_color=color, fill_type='solid')
    halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin   = Side(style='thin', color='FFFFFF')
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font      = hfont
        cell.fill      = hfill
        cell.alignment = halign
        cell.border    = bdr
    ws.row_dimensions[1].height = 32


def _style_data(ws, num_rows: int, num_cols: int):
    even_fill = PatternFill(start_color=COLORS['row_even'], end_color=COLORS['row_even'], fill_type='solid')
    thin  = Side(style='thin', color='D5D8DC')
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    align = Alignment(vertical='center', wrap_text=False)
    for r in range(2, num_rows + 2):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border    = bdr
            cell.alignment = align
            if r % 2 == 0:
                cell.fill = even_fill


def _add_total_row(ws, df: pd.DataFrame, num_cols: int):
    """Add a bold TOTAL row at the bottom for numeric columns."""
    total_fill = PatternFill(start_color=COLORS['total_row'], end_color=COLORS['total_row'], fill_type='solid')
    tfont      = Font(name='Calibri', bold=True, size=11)
    num_row    = len(df) + 2  # header=1, data rows=len(df), total=next

    ws.cell(row=num_row, column=1).value = 'TOTAL'
    ws.cell(row=num_row, column=1).font  = tfont

    numeric_cols = df.select_dtypes(include=['float64', 'int64']).columns
    for col_name in numeric_cols:
        col_idx = list(df.columns).index(col_name) + 1
        col_sum = df[col_name].sum()
        cell = ws.cell(row=num_row, column=col_idx)
        cell.value  = round(float(col_sum), 2)
        cell.font   = tfont

    for c in range(1, num_cols + 1):
        cell = ws.cell(row=num_row, column=c)
        cell.fill = total_fill
        thin = Side(style='medium', color='1A5276')
        cell.border = Border(top=thin, bottom=thin)


def _format_numbers(ws, df: pd.DataFrame):
    currency = '#,##0.00'
    integer  = '#,##0'
    numeric_cols = df.select_dtypes(include=['float64', 'int64']).columns
    for col in numeric_cols:
        idx = list(df.columns).index(col) + 1
        fmt = integer if any(w in col for w in ['Spots', 'Duration']) else currency
        for r in range(2, len(df) + 3):   # +3 to include total row
            ws.cell(row=r, column=idx).number_format = fmt


def _auto_width(ws, df: pd.DataFrame, min_w=10, max_w=42):
    for i, col in enumerate(df.columns, 1):
        h_len    = len(str(col))
        data_len = int(df[col].astype(str).str.len().max()) if len(df) > 0 else 0
        width    = max(min_w, min(max(h_len, data_len) + 3, max_w))
        ws.column_dimensions[get_column_letter(i)].width = width


def _freeze_filter(ws, df: pd.DataFrame):
    ws.freeze_panes = 'A2'
    last_col = get_column_letter(len(df.columns))
    last_row = len(df) + 1
    ws.auto_filter.ref = f'A1:{last_col}{last_row}'


# ─────────────────────────────────────────────────────────────────
# BROADCASTER — Build spot-level data with 18 columns
# ─────────────────────────────────────────────────────────────────

BROADCASTER_COLUMNS = [
    'Advertiser Name',
    'Broadcaster Name',
    'Agency Name',
    'Channel Name/STN',
    'Billing Period',
    'PO Number',
    'Invoice Number',
    'Invoice Date',
    'TP (Telecast Program)',
    'Program',
    'Date',
    'Day',
    'Air Time',
    'Duration (Sec)',
    'Spot Copy (Caption)',
    'Brand',
    'Rate (INR)',
    'Amount (INR)',
]


def build_broadcaster_data(parsed_invoices: List[ParsedBroadcasterInvoice]) -> pd.DataFrame:
    """Build individual spot-level DataFrame from broadcaster invoices.

    Each row = one individual spot with all 18 fields:
    Advertiser, Broadcaster, Agency, Channel/STN, Billing Period, PO,
    Invoice No, Invoice Date, TP, Program, Date, Day, Air Time,
    Duration, Spot Copy, Brand, Rate, Amount.
    """
    rows = []

    for inv in parsed_invoices:
        h = inv.header

        for spot in inv.spots:
            rows.append({
                'Advertiser Name':          h.advertiser_name,
                'Broadcaster Name':         h.broadcaster_name,
                'Agency Name':              h.agency_name,
                'Channel Name/STN':         h.channel_name or '',
                'Billing Period':           h.billing_period,
                'PO Number':                h.po_number,
                'Invoice Number':           h.invoice_number,
                'Invoice Date':             h.invoice_date,
                'TP (Telecast Program)':    spot.tp,
                'Program':                  spot.program,
                'Date':                     spot.date,
                'Day':                      spot.day,
                'Air Time':                 spot.air_time,
                'Duration (Sec)':           spot.duration,
                'Spot Copy (Caption)':      spot.spot_copy,
                'Brand':                    spot.brand or h.brand,
                'Rate (INR)':               spot.rate,
                'Amount (INR)':             spot.amount,
            })

    if not rows:
        return pd.DataFrame(columns=BROADCASTER_COLUMNS)

    df = pd.DataFrame(rows, columns=BROADCASTER_COLUMNS)
    df = df.sort_values(['Invoice Number', 'Channel Name/STN', 'Date']).reset_index(drop=True)
    return df


def broadcaster_spots_to_dicts(inv: ParsedBroadcasterInvoice) -> List[Dict]:
    """Used by app.py for UI preview — individual spot-level rows with all 18 fields."""
    h = inv.header
    rows = []
    for spot in inv.spots:
        rows.append({
            'Invoice Number':           h.invoice_number,
            'Channel Name':             h.channel_name,
            'TP (Telecast Program)':    spot.tp,
            'Program':                  spot.program,
            'Date':                     spot.date,
            'Day':                      spot.day,
            'Air Time':                 spot.air_time,
            'Duration (Sec)':           spot.duration,
            'Spot Copy (Caption)':      spot.spot_copy,
            'Brand':                    spot.brand or h.brand,
            'Rate (INR)':               spot.rate,
            'Amount (INR)':             spot.amount,
        })
    return rows


def broadcaster_to_summary_dict(inv: ParsedBroadcasterInvoice) -> Dict:
    """Create a summary dict for app.py preview."""
    h = inv.header
    return {
        'Invoice Number':       h.invoice_number,
        'Invoice Date':         h.invoice_date,
        'Broadcaster Name':     h.broadcaster_name,
        'Advertiser Name':      h.advertiser_name,
        'Channel Name':         h.channel_name,
        'Brand':                h.brand,
        'Billing Period':       h.billing_period,
        'PO Number':            h.po_number,
        'Total Spots':          len(inv.spots),
        'Net Amount':           round(h.net_amount, 2),
        'Format':               inv.format_type,
        'Source File':          inv.source_file,
        'Errors':               '; '.join(inv.errors) if inv.errors else '',
    }


# ─────────────────────────────────────────────────────────────────
# SHEET 0 — Invoice Summary Dashboard (New!)
# ─────────────────────────────────────────────────────────────────

INVOICE_SUMMARY_COLUMNS = [
    'Invoice Type',
    'Invoice Number',
    'Invoice Date',
    'Broadcaster / Channel Name',
    'Agency Name',
    'Advertiser Name',
    'Brand Name',
    'PO Number',
    'Total Spots',
    'Total Net Amount (INR)',
    'Total Amount Payable (INR)',
    'Billing Period',
    'Source File Name',
]

def build_invoice_summary_sheet(parsed_agency: List[ParsedInvoice], parsed_broadcaster: List[ParsedBroadcasterInvoice]) -> pd.DataFrame:
    rows = []
    
    # Agency Invoices
    for r in parsed_agency:
        h = r.header
        rows.append({
            'Invoice Type': 'Agency',
            'Invoice Number': h.invoice_number,
            'Invoice Date': h.invoice_date,
            'Broadcaster / Channel Name': 'Multiple (Agency)',
            'Agency Name': h.agency_name,
            'Advertiser Name': h.advertiser_name,
            'Brand Name': h.brand_name,
            'PO Number': h.po_number,
            'Total Spots': r.total_spots if r.total_spots > 0 else len(r.spot_details),
            'Total Net Amount (INR)': _parse_number(h.net_cost),
            'Total Amount Payable (INR)': _parse_number(h.total_amount_payable),
            'Billing Period': h.activity_month,
            'Source File Name': getattr(r, 'source_file', 'Agency Invoice.pdf'),
        })
        
    # Broadcaster Invoices
    for r in parsed_broadcaster:
        h = r.header
        rows.append({
            'Invoice Type': 'Broadcaster (%s)' % r.format_type.upper(),
            'Invoice Number': h.invoice_number,
            'Invoice Date': h.invoice_date,
            'Broadcaster / Channel Name': h.channel_name or h.broadcaster_name,
            'Agency Name': h.agency_name,
            'Advertiser Name': h.advertiser_name,
            'Brand Name': h.brand,
            'PO Number': h.po_number,
            'Total Spots': len(r.spots),
            'Total Net Amount (INR)': round(h.net_amount, 2),
            'Total Amount Payable (INR)': round(h.net_amount, 2),
            'Billing Period': h.billing_period,
            'Source File Name': r.source_file,
        })
        
    df = pd.DataFrame(rows, columns=INVOICE_SUMMARY_COLUMNS)
    return df


# ─────────────────────────────────────────────────────────────────
# Main Export — handles both Agency and Broadcaster
# ─────────────────────────────────────────────────────────────────

def export_to_excel(dfs: Dict[str, pd.DataFrame], output_path: str) -> str:
    """
    Export clean, readable Excel. Supports:
    - 'summary' key → Invoice Summary sheet
    - 'channel' key → Agency Details sheet
    - 'broadcaster' key → Broadcaster Details sheet
    """
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        sheets_written = 0

        # High-level summary sheet
        df_sum = dfs.get('summary', pd.DataFrame())
        if df_sum is not None and len(df_sum) > 0:
            df_sum.to_excel(writer, sheet_name='Invoice Summary', index=False)
            ws = writer.sheets['Invoice Summary']
            _style_header(ws, len(df_sum.columns), COLORS['sheet3_header'])
            _style_data(ws, len(df_sum), len(df_sum.columns))
            _add_total_row(ws, df_sum, len(df_sum.columns))
            _format_numbers(ws, df_sum)
            _auto_width(ws, df_sum)
            _freeze_filter(ws, df_sum)
            ws.sheet_properties.tabColor = '1A5276'
            ws.sheet_view.showGridLines  = False
            sheets_written += 1

        # Agency details sheet
        df_agency = dfs.get('channel', pd.DataFrame())
        if df_agency is not None and len(df_agency) > 0:
            sheet_name = 'Agency Details'
            df_agency.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            _style_header(ws, len(df_agency.columns), COLORS['sheet1_header'])
            _style_data(ws, len(df_agency), len(df_agency.columns))
            _add_total_row(ws, df_agency, len(df_agency.columns))
            _format_numbers(ws, df_agency)
            _auto_width(ws, df_agency)
            _freeze_filter(ws, df_agency)
            ws.sheet_properties.tabColor = '1F3864'
            ws.sheet_view.showGridLines  = False
            sheets_written += 1

        # Broadcaster details sheet
        df_bc = dfs.get('broadcaster', pd.DataFrame())
        if df_bc is not None and len(df_bc) > 0:
            sheet_name = 'Broadcaster Details'
            df_bc.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            _style_header(ws, len(df_bc.columns), COLORS['sheet2_header'])
            _style_data(ws, len(df_bc), len(df_bc.columns))
            _add_total_row(ws, df_bc, len(df_bc.columns))
            _format_numbers(ws, df_bc)
            _auto_width(ws, df_bc)
            _freeze_filter(ws, df_bc)
            ws.sheet_properties.tabColor = '5B2C6F'
            ws.sheet_view.showGridLines  = False
            sheets_written += 1

        # Fallback: if no data sheets written, create an info sheet
        if sheets_written == 0:
            info_df = pd.DataFrame({
                'Message': ['No invoice data could be extracted from the uploaded PDF(s).'],
                'Suggestion': ['Please check the PDF format and try again.']
            })
            info_df.to_excel(writer, sheet_name='Info', index=False)

    agency_rows = len(df_agency) if df_agency is not None and len(df_agency) > 0 else 0
    bc_rows     = len(df_bc) if df_bc is not None and len(df_bc) > 0 else 0
    logger.info('Excel exported: %s | Agency=%d rows | Broadcaster=%d rows', output_path, agency_rows, bc_rows)
    return output_path


# ─────────────────────────────────────────────────────────────────
# Helpers for app.py
# ─────────────────────────────────────────────────────────────────

def invoice_to_summary_dict(result: ParsedInvoice) -> Dict:
    h = result.header
    return {
        'Invoice Number':       h.invoice_number,
        'Invoice Date':         h.invoice_date,
        'Agency Name':          h.agency_name,
        'Advertiser Name':      h.advertiser_name,
        'Brand Name':           h.brand_name,
        'Campaign Name':        h.campaign_name,
        'Activity Month':       h.activity_month,
        'Estimate Number':      h.estimate_number,
        'Estimate Period':      h.estimate_period,
        'PO Number':            h.po_number,
        'Total Amount Payable': _parse_number(h.total_amount_payable),
        'Total Spots':          result.total_spots if result.total_spots > 0 else len(result.spot_details),
        'Errors':               '; '.join(result.errors) if result.errors else '',
    }


def spots_to_dicts(result: ParsedInvoice) -> List[Dict]:
    """Used by app.py for UI preview (channel-wise aggregated)."""
    ann1: Dict[tuple, str] = {}
    for ar in result.annexure_records:
        ann1[(ar.channel.strip(), ar.program.strip())] = ar.dates_with_spots

    groups: Dict[tuple, dict] = {}
    for sd in result.spot_details:
        key = (sd.channel, sd.program, sd.time_band, sd.producer, sd.spot_duration, sd.net_spot_rate_per_10sec)
        if key not in groups:
            groups[key] = {'count': 0, 'net_cost': 0.0}
        groups[key]['count']    += 1
        groups[key]['net_cost'] += sd.net_cost

    rows = []
    for key, agg in groups.items():
        channel, program, time_band, producer, duration, rate = key
        rows.append({
            'Invoice Number':              result.header.invoice_number,
            'Channel Name':                channel,
            'Program':                     program,
            'Time Band':                   time_band,
            'Broadcaster Name (Producer)': producer,
            'Spot Duration (Sec)':         duration,
            'Spot Rate (Per 10 Sec)':      rate,
            'No. of Spots':                agg['count'],
            'Total Net Cost':              round(agg['net_cost'], 2),
            'Date Wise Spots':             ann1.get((channel.strip(), program.strip()), ''),
        })
    return rows


def get_results_summary(parsed_invoices: List[ParsedInvoice]) -> Dict:
    total_spots       = 0
    total_net_cost    = 0.0
    total_amt_payable = 0.0
    for r in parsed_invoices:
        total_spots       += r.total_spots if r.total_spots > 0 else len(r.spot_details)
        total_net_cost    += _parse_number(r.header.net_cost)
        total_amt_payable += _parse_number(r.header.total_amount_payable)
    return {
        'total_invoices':       len(parsed_invoices),
        'total_spots':          total_spots,
        'total_net_cost':       round(total_net_cost, 2),
        'total_amount_payable': round(total_amt_payable, 2),
    }


# ─────────────────────────────────────────────────────────────────
# Quick test
# ─────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    import glob
    invoice_dir = r'C:\Users\asus\Downloads\Sustenance_Redmi_Note_13_May24\Sustenance_Redmi_Note_13_May24\Agency Invoices'
    pdf_files   = sorted(glob.glob(os.path.join(invoice_dir, '*.pdf')))
    print('Found %d PDFs' % len(pdf_files))

    parsed = [parse_invoice(p) for p in pdf_files]
    dfs    = process_multiple_invoices(parsed)

    out = os.path.join('.', 'output', 'Agency_Invoices_Clean.xlsx')
    os.makedirs('output', exist_ok=True)
    export_to_excel(dfs, out)

    print('Excel saved:', out)
    print('Sheet 1 — Invoice Summary :', len(dfs['summary']), 'rows')
    print('Sheet 2 — Channel-wise    :', len(dfs['channel']), 'rows')
    print('(Raw spots for reference  :', len(dfs['main']), 'spots)')
